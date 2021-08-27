import openpyxl 


xlsx_file = './excel-sample.xlsx'
wb_obj = openpyxl.load_workbook(xlsx_file) 
sheet = wb_obj.active

col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
   
print(col_names)
print(sheet.cell(row=3, column=2).value)
print(sheet.max_row, sheet.max_column)
